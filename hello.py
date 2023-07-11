from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import Tk, StringVar,OptionMenu, Button, Label, Entry
import barcode
from tkcalendar import DateEntry
from barcode.writer import ImageWriter
from PIL import Image, ImageDraw, ImageFont


def read_barcodes_from_excel(file_path, sheet_name):
    workbook= load_workbook(file_path)
    sheet=workbook[sheet_name]
    barcodes=["                       "]
    for row in sheet.iter_rows(values_only=True):
        barcode_value=row[0]
        barcodes.append(barcode_value)
    return barcodes

def generate_barcode(selected_barcode):
    selected_date = date_entry.get_date().strftime("%Y-%m-%d")
    Code128= barcode.get_barcode_class('code128')
    code128= Code128(selected_barcode, writer=ImageWriter())
    barcode_image= code128.render()

    draw= ImageDraw.Draw(barcode_image)
    font= ImageFont.load_default()
    date_text= selected_date
    text_bbox= draw.textbbox((0,0),date_text, font=font)
    x = (barcode_image.width - text_bbox[2]) // 2
    y = barcode_image.height - text_bbox[3] - 10
    draw.text((x, y), date_text, font=font, fill="black")

    return barcode_image


def generate_selected_barcode():
    selected_barcode = selected_barcode_var.get()
    quantity= int(quantity_entry.get())

    for i in range(quantity):
        barcode_image= generate_barcode(selected_barcode)
        barcode_image.save(f'barcode_{selected_barcode}_{i + 1}.png')

root=Tk()
root.title('Barcode Dropdown')
root.geometry('1200x1200')


barcodes = read_barcodes_from_excel('data_barcodes.xlsx', 'Sheet1')

selected_barcode_var = StringVar(root)
selected_barcode_var.set(barcodes[0])

selected_barcode_label= Label(root, text="Select Barcode", font=('Helvatical Bold',15))
selected_barcode_label.grid(row=0, column=0, sticky='w', padx=20, pady=10)

barcode_dropdown = OptionMenu(root, selected_barcode_var, *barcodes)
barcode_dropdown.grid(row=0, column=1, padx=20, pady=10)

date_label = Label(root, text="Select Date:")
date_label.grid(row=1, column=0, sticky="w", padx=20, pady=5)

date_entry = DateEntry(root, date_pattern="yyyy-mm-dd")
date_entry.grid(row=1, column=1, padx=20, pady=5)


quantity_label= Label(root, text="Enter Quantity")
quantity_label.grid(row=2, column=0, sticky="w", padx=20, pady=5)

quantity_entry = Entry(root)
quantity_entry.grid(row=2, column=1, padx=20, pady=5)

generate_button= Button(root, text="Generate Barcode", command= generate_selected_barcode)
generate_button.grid(row=3, column=0, columnspan=2, padx=20, pady=20)



root.mainloop()
